from __future__ import annotations

import json
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from numbers import Number
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .query_executor import QueryPayload


SUMMARY_SHEET = "_summary"
CHECKPOINT_HEADERS = [
    "Date",
    "Status",
    "DurationSeconds",
    "PayloadJSON",
    "RenderedQuery",
    "ErrorMessage",
    "UpdatedAtUTC",
]


@dataclass(frozen=True)
class DayWindow:
    day_key: str
    start_iso: str
    end_iso: str


@dataclass(frozen=True)
class DailyCheckpointRecord:
    day_key: str
    status: str
    duration_seconds: float
    payload: QueryPayload | None
    rendered_query: str
    error_message: str


def parse_utc_datetime(value: str) -> datetime:
    text = value.strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    parsed = datetime.fromisoformat(text)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    return parsed.astimezone(UTC)


def format_utc_datetime(value: datetime) -> str:
    return value.astimezone(UTC).isoformat().replace("+00:00", "Z")


def iter_day_windows(start_date_text: str, end_date_text: str) -> list[DayWindow]:
    start_at = parse_utc_datetime(start_date_text)
    end_at = parse_utc_datetime(end_date_text)
    if end_at <= start_at:
        raise ValueError(
            f"Invalid day-wise date window: end_date ({end_date_text}) must be greater than start_date ({start_date_text})."
        )

    current = start_at
    day_windows: list[DayWindow] = []
    while current < end_at:
        next_day = min(current + timedelta(days=1), end_at)
        day_windows.append(
            DayWindow(
                day_key=current.date().isoformat(),
                start_iso=format_utc_datetime(current),
                end_iso=format_utc_datetime(next_day),
            )
        )
        current = next_day
    return day_windows


def _is_numeric(value: Any) -> bool:
    return isinstance(value, Number) and not isinstance(value, bool)


def summarize_payload(payload: QueryPayload) -> str:
    if isinstance(payload, dict):
        keys = ", ".join(sorted(payload.keys()))
        return f"scalar keys=[{keys}]"
    return f"rows={len(payload)}"


def aggregate_scalar_payloads(payloads: Iterable[dict[str, Any]]) -> dict[str, Any]:
    aggregated: dict[str, Any] = {}
    for payload in payloads:
        for key, value in payload.items():
            if _is_numeric(value):
                existing = aggregated.get(key)
                aggregated[key] = value if existing is None else existing + value
            elif key not in aggregated or aggregated[key] in (None, ""):
                aggregated[key] = value
    return aggregated


def aggregate_table_payloads(payloads: Iterable[list[dict[str, Any]]], columns: list[str]) -> list[dict[str, Any]]:
    payload_list = list(payloads)
    numeric_columns = {
        column
        for column in columns
        if any(_is_numeric(row.get(column)) for payload in payload_list for row in payload)
    }
    dimension_columns = [column for column in columns if column not in numeric_columns]

    aggregated: dict[tuple[Any, ...], dict[str, Any]] = {}
    for payload in payload_list:
        for row in payload:
            row_key = tuple(row.get(column) for column in dimension_columns)
            if row_key not in aggregated:
                aggregated[row_key] = {column: row.get(column) for column in columns}
                for numeric_column in numeric_columns:
                    value = row.get(numeric_column)
                    aggregated[row_key][numeric_column] = value if _is_numeric(value) else 0
                continue

            target = aggregated[row_key]
            for numeric_column in numeric_columns:
                value = row.get(numeric_column)
                if _is_numeric(value):
                    target[numeric_column] = target.get(numeric_column, 0) + value
            for dimension_column in dimension_columns:
                if target.get(dimension_column) in (None, ""):
                    target[dimension_column] = row.get(dimension_column)

    return list(aggregated.values())


class DailyCheckpointStore:
    def __init__(self, output_path: Path) -> None:
        self.output_path = output_path

    def load_successful_records(self, query_id: str) -> dict[str, DailyCheckpointRecord]:
        workbook = self._load_or_create_workbook()
        worksheet = self._get_query_sheet(workbook, query_id, create=False)
        if worksheet is None:
            return {}

        records: dict[str, DailyCheckpointRecord] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            day_key = str(row[0] or "").strip()
            status = str(row[1] or "").strip()
            if not day_key or status != "success":
                continue

            payload_raw = str(row[3] or "").strip()
            payload = json.loads(payload_raw) if payload_raw else None
            rendered_query = str(row[4] or "")
            duration_seconds = float(row[2] or 0)
            error_message = str(row[5] or "")
            records[day_key] = DailyCheckpointRecord(
                day_key=day_key,
                status=status,
                duration_seconds=duration_seconds,
                payload=payload,
                rendered_query=rendered_query,
                error_message=error_message,
            )
        return records

    def save_success(
        self,
        query_id: str,
        day_key: str,
        payload: QueryPayload,
        rendered_query: str,
        duration_seconds: float,
    ) -> None:
        self._upsert_record(
            query_id=query_id,
            day_key=day_key,
            status="success",
            duration_seconds=duration_seconds,
            payload=payload,
            rendered_query=rendered_query,
            error_message="",
        )

    def save_failure(
        self,
        query_id: str,
        day_key: str,
        rendered_query: str,
        error_message: str,
    ) -> None:
        self._upsert_record(
            query_id=query_id,
            day_key=day_key,
            status="failed",
            duration_seconds=0.0,
            payload=None,
            rendered_query=rendered_query,
            error_message=error_message,
        )

    def _upsert_record(
        self,
        query_id: str,
        day_key: str,
        status: str,
        duration_seconds: float,
        payload: QueryPayload | None,
        rendered_query: str,
        error_message: str,
    ) -> None:
        workbook = self._load_or_create_workbook()
        worksheet = self._get_query_sheet(workbook, query_id, create=True)
        assert worksheet is not None

        target_row = None
        for row_index in range(2, worksheet.max_row + 1):
            if str(worksheet.cell(row=row_index, column=1).value or "").strip() == day_key:
                target_row = row_index
                break
        if target_row is None:
            target_row = worksheet.max_row + 1

        payload_json = json.dumps(payload, ensure_ascii=True, default=str) if payload is not None else ""
        updated_at = format_utc_datetime(datetime.now(tz=UTC))
        values = [day_key, status, duration_seconds, payload_json, rendered_query, error_message, updated_at]
        for column_index, value in enumerate(values, start=1):
            worksheet.cell(row=target_row, column=column_index, value=value)

        workbook.save(self.output_path)

    def _load_or_create_workbook(self) -> Workbook:
        if self.output_path.exists():
            return load_workbook(self.output_path)

        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.active.title = SUMMARY_SHEET
        workbook.active.append(["Daily query checkpoints"])
        workbook.save(self.output_path)
        return workbook

    def _get_query_sheet(self, workbook: Workbook, query_id: str, create: bool) -> Any:
        if query_id in workbook.sheetnames:
            return workbook[query_id]
        if not create:
            return None

        worksheet = workbook.create_sheet(title=query_id)
        worksheet.append(CHECKPOINT_HEADERS)
        return worksheet
