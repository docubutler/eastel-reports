from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


SCALAR_PLACEHOLDER_PATTERN = re.compile(r"%(Q\d{3})\.([A-Za-z0-9_]+)%")
TABLE_PLACEHOLDER_PATTERN = re.compile(r"%TABLE:(Q\d{3})%")
HEADER_FILL = PatternFill(fill_type="solid", start_color="4F81BD", end_color="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF")


@dataclass(frozen=True)
class ScalarPlaceholder:
    sheet_name: str
    cell_reference: str
    query_id: str
    field_name: str
    raw_value: str


@dataclass(frozen=True)
class TableAnchor:
    sheet_name: str
    cell_reference: str
    query_id: str
    anchor_text: str


@dataclass(frozen=True)
class WorkbookReferences:
    scalar_placeholders: list[ScalarPlaceholder]
    table_anchors: list[TableAnchor]

    @property
    def query_ids(self) -> list[str]:
        ordered: list[str] = []
        for query_id in [*self.scalar_query_ids, *self.table_query_ids]:
            if query_id not in ordered:
                ordered.append(query_id)
        return ordered

    @property
    def scalar_query_ids(self) -> list[str]:
        ordered: list[str] = []
        for placeholder in self.scalar_placeholders:
            if placeholder.query_id not in ordered:
                ordered.append(placeholder.query_id)
        return ordered

    @property
    def table_query_ids(self) -> list[str]:
        ordered: list[str] = []
        for anchor in self.table_anchors:
            if anchor.query_id not in ordered:
                ordered.append(anchor.query_id)
        return ordered


def _format_cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    return value


def scan_workbook_references(template_path: Path) -> WorkbookReferences:
    workbook = load_workbook(template_path)
    scalar_placeholders: list[ScalarPlaceholder] = []
    table_anchors: list[TableAnchor] = []

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue

                raw_value = cell.value
                for match in SCALAR_PLACEHOLDER_PATTERN.finditer(raw_value):
                    scalar_placeholders.append(
                        ScalarPlaceholder(
                            sheet_name=worksheet.title,
                            cell_reference=cell.coordinate,
                            query_id=match.group(1),
                            field_name=match.group(2),
                            raw_value=raw_value,
                        )
                    )

                table_match = TABLE_PLACEHOLDER_PATTERN.fullmatch(raw_value.strip())
                if table_match:
                    table_anchors.append(
                        TableAnchor(
                            sheet_name=worksheet.title,
                            cell_reference=cell.coordinate,
                            query_id=table_match.group(1),
                            anchor_text=raw_value.strip(),
                        )
                    )

    return WorkbookReferences(
        scalar_placeholders=scalar_placeholders,
        table_anchors=table_anchors,
    )


def _copy_row_style(worksheet: Worksheet, source_row: int, target_row: int) -> None:
    for column_index in range(1, worksheet.max_column + 1):
        source_cell = worksheet.cell(row=source_row, column=column_index)
        target_cell = worksheet.cell(row=target_row, column=column_index)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height


def _ensure_header_row(
    worksheet: Worksheet,
    anchor_row: int,
    start_column: int,
    columns: list[str],
) -> int:
    header_row = anchor_row - 1
    if header_row < 1:
        worksheet.insert_rows(anchor_row)
        header_row = anchor_row
        anchor_row += 1
    else:
        existing_values = [
            worksheet.cell(row=header_row, column=start_column + offset).value
            for offset in range(len(columns))
        ]
        if any(value not in (None, "") for value in existing_values):
            return anchor_row

        worksheet.insert_rows(anchor_row)
        header_row = anchor_row
        anchor_row += 1

    for offset, column_name in enumerate(columns):
        header_cell = worksheet.cell(row=header_row, column=start_column + offset)
        header_cell.value = column_name
        header_cell.fill = copy(HEADER_FILL)
        header_cell.font = copy(HEADER_FONT)

    return anchor_row


def _replace_scalar_placeholders(worksheet: Worksheet, scalar_results: dict[str, dict[str, Any]]) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue

            raw_value = cell.value

            def replace(match: re.Match[str]) -> str:
                query_id = match.group(1)
                field_name = match.group(2)
                result = scalar_results.get(query_id)
                if result is None:
                    raise ValueError(f"Missing scalar result for query {query_id}.")
                if field_name not in result:
                    raise ValueError(f"Scalar result for query {query_id} does not contain field '{field_name}'.")
                value = _format_cell_value(result[field_name])
                return "" if value is None else str(value)

            replaced_value = SCALAR_PLACEHOLDER_PATTERN.sub(replace, raw_value)
            if replaced_value != raw_value:
                cell.value = replaced_value


def _render_table_anchor(
    worksheet: Worksheet,
    anchor_row: int,
    anchor_column: int,
    rows: list[dict[str, Any]],
    columns: list[str],
) -> None:
    anchor_row = _ensure_header_row(worksheet, anchor_row, anchor_column, columns)
    template_row = anchor_row

    if rows:
        worksheet.insert_rows(anchor_row, amount=len(rows))
        for row_offset, row_payload in enumerate(rows):
            target_row = anchor_row + row_offset
            _copy_row_style(worksheet, template_row + len(rows), target_row)
            for column_offset, column_name in enumerate(columns):
                target_cell = worksheet.cell(row=target_row, column=anchor_column + column_offset)
                target_cell.value = _format_cell_value(row_payload.get(column_name))

    worksheet.delete_rows(template_row + len(rows), 1)


def render_report_workbook(
    template_path: Path,
    output_path: Path,
    scalar_results: dict[str, dict[str, Any]],
    table_results: dict[str, list[dict[str, Any]]],
    table_definitions: dict[str, list[str]],
) -> None:
    workbook = load_workbook(template_path)

    for worksheet in workbook.worksheets:
        _replace_scalar_placeholders(worksheet, scalar_results)

        anchors: list[tuple[int, int, str]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    match = TABLE_PLACEHOLDER_PATTERN.fullmatch(cell.value.strip())
                    if match:
                        anchors.append((cell.row, cell.column, match.group(1)))

        for anchor_row, anchor_column, query_id in sorted(anchors, key=lambda item: item[0], reverse=True):
            if query_id not in table_definitions:
                raise ValueError(f"Missing table column definition for query {query_id}.")
            _render_table_anchor(
                worksheet=worksheet,
                anchor_row=anchor_row,
                anchor_column=anchor_column,
                rows=table_results.get(query_id, []),
                columns=table_definitions[query_id],
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def apply_scalar_query_result(
    workbook_path: Path,
    query_id: str,
    scalar_result: dict[str, Any],
) -> int:
    workbook = load_workbook(workbook_path)
    replacement_count = 0

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue

                raw_value = cell.value
                matched = False

                def replace(match: re.Match[str]) -> str:
                    nonlocal matched
                    placeholder_query_id = match.group(1)
                    field_name = match.group(2)
                    if placeholder_query_id != query_id:
                        return match.group(0)
                    if field_name not in scalar_result:
                        raise ValueError(
                            f"Scalar result for query {query_id} does not contain field '{field_name}'."
                        )
                    matched = True
                    value = _format_cell_value(scalar_result[field_name])
                    return "" if value is None else str(value)

                replaced_value = SCALAR_PLACEHOLDER_PATTERN.sub(replace, raw_value)
                if matched:
                    cell.value = replaced_value
                    replacement_count += 1

    workbook.save(workbook_path)
    return replacement_count


def apply_table_query_result(
    workbook_path: Path,
    query_id: str,
    rows: list[dict[str, Any]],
    columns: list[str],
) -> int:
    workbook = load_workbook(workbook_path)
    anchor_count = 0

    for worksheet in workbook.worksheets:
        anchors: list[tuple[int, int]] = []
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                match = TABLE_PLACEHOLDER_PATTERN.fullmatch(cell.value.strip())
                if match and match.group(1) == query_id:
                    anchors.append((cell.row, cell.column))

        for anchor_row, anchor_column in sorted(anchors, key=lambda item: item[0], reverse=True):
            _render_table_anchor(
                worksheet=worksheet,
                anchor_row=anchor_row,
                anchor_column=anchor_column,
                rows=rows,
                columns=columns,
            )
            anchor_count += 1

    workbook.save(workbook_path)
    return anchor_count


def write_actual_queries_workbook(output_path: Path, rendered_queries: list[tuple[str, str, str]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Query Definitions"
    worksheet.append(["Query ID", "Query File", "Rendered Query"])

    for row in rendered_queries:
        worksheet.append(list(row))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
